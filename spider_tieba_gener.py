'''
爬取XX吧首页

'''
import re
import time
from urllib import request,parse
import openpyxl
from bs4 import BeautifulSoup
import os

class Spider:
    def __init__(self):
        self.headers = {
            "User-Agent":
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36",
            "Accept":
                "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9"
        }

    def getContent(self,url):
        '''
        获取页面源代码
        :param url:
        :return:
        '''

        response = request.Request(url,headers=self.headers)
        r = request.urlopen(response)
        html = r.read().decode("UTF-8")
        r.close()
        return html


    def saveHTML(self,html):
        '''
        将html保存
        debug用
        :param html: 网页文本
        :return:
        '''
        f = open("src.html", "w+", encoding="UTF-8")
        f.write(html)
        f.close()


    def getInfo(self,html,items,*temps):
        '''
        获取帖子信息（包含评论数、标题、楼层地址、大致内容、楼主、时间）
        并将其返回
        :param html: 网页源代码
        :param items: 保存位置，列表
        :param temps: 元组形式，传入时间，内容分别为年月日
        :return: items
        '''

        # 通过正则表达式匹配指定信息
        findItem = re.compile('''(<li class=" j_thread_list.*?clearfix thread_item_box".*?>.*?</li>)''',re.S)

        for item in re.findall(findItem,html):
            item = BeautifulSoup(item,"html.parser")
            title = item.find('a',attrs={"class": "j_th_tit"})
            # 获取楼层地址和标题名
            topicUrl = title['href']
            topicName = title.string
            # 获取评论数
            num = item.find("span",attrs={"class": "threadlist_rep_num center_text"}).string
            # 获取大致内容
            content = item.find("div",attrs={"class": "threadlist_abs"})
            if content is not None:
                conFinal = content.text.strip()
            else:
                conFinal = ""

            # 获取楼主名字
            author = item.find("a",attrs={"class": "frs-author-name"}).text.strip()
            # 获取时间，并对时间进行操作
            t = item.find("span",attrs={"class": "threadlist_reply_date"})
            if t is not None:
                t_f = t.text.strip()

                if t_f.__contains__(':'):
                    # 判断如果为当天时间（hh:mm）,则加上年月日
                    t_f = temps[0]+"-"+temps[1]+"-"+temps[2]+" "+t_f
                elif t_f.__contains__("-"):
                    if len(t_f) > 5:
                        # 如果是往年的时间，则直接不用做操作（长度大于5）
                        pass
                    else:
                        # 判断如果为今年的非当天时间（MM-dd）,则加上年
                        t_f = temps[0] + "-" + t_f
            else:
                t_f = ""

            items.append([num,topicName,topicUrl,conFinal,author,t_f])
        return items

    def saveData(self, data, name, num):
        file = name + str(num) + ".xlsx"
        dir = name + "吧"

        # 创建目录存放数据，如果目录已经存在则不创建
        if os.path.exists(dir):
            pass
        else:
            os.mkdir(dir)

        try:
            # 判断xls文件是否存在，不存在则创建，存在则追加数据
            if os.path.exists(file):
                wb_write = openpyxl.load_workbook(dir+"/"+file)
                ws_write = wb_write[name]

                nrows = ws_write.max_row  # 行数
                ncols = ws_write.max_column  # 列数
                for i in range(len(data)):
                    ws_write.cell(i + 1 + nrows, 1, data[i][1])  # [1]
                    ws_write.cell(i + 1 + nrows, 2, data[i][0])  # [0]
                    ws_write.cell(i + 1 + nrows, 3, data[i][2])  # [2]
                    ws_write.cell(i + 1 + nrows, 4, data[i][3])  # [3]
                    ws_write.cell(i + 1 + nrows, 5, data[i][4])  # [4]
                    ws_write.cell(i + 1 + nrows, 6, data[i][5])  # [5]
                wb_write.save(dir+"/"+file)
            else:
                wb = openpyxl.Workbook()
                ws = wb.create_sheet(name)

                ws.cell(1, 1, "楼层名称")  # [1]
                ws.cell(1, 2, "评论数")  # [0]
                ws.cell(1, 3, "楼层地址")  # [2]
                ws.cell(1, 4, "大致内容")  # [3]
                ws.cell(1, 5, "楼主")  # [4]
                ws.cell(1, 6, "时间")  # [5]
                for i in range(len(data)):
                    ws.cell(i + 2, 1, data[i][1])  # [1]
                    ws.cell(i + 2, 2, data[i][0])  # [0]
                    ws.cell(i + 2, 3, data[i][2])  # [2]
                    ws.cell(i + 2, 4, data[i][3])  # [3]
                    ws.cell(i + 2, 5, data[i][4])  # [4]
                    ws.cell(i + 2, 6, data[i][5])  # [5]

                wb.save(dir+"/"+file)
        except ValueError:
            wb_new = openpyxl.Workbook()
            ws_new = wb_new.create_sheet(name)

            ws_new.cell(1, 1, "楼层名称")  # [1]
            ws_new.cell(1, 2, "评论数")  # [0]
            ws_new.cell(1, 3, "楼层地址")  # [2]
            ws_new.cell(1, 4, "大致内容")  # [3]
            ws_new.cell(1, 5, "楼主")  # [4]
            ws_new.cell(1, 6, "时间")  # [5]
            for i in range(len(data)):
                ws_new.cell(i + 2, 1, data[i][1])  # [1]
                ws_new.cell(i + 2, 2, data[i][0])  # [0]
                ws_new.cell(i + 2, 3, data[i][2])  # [2]
                ws_new.cell(i + 2, 4, data[i][3])  # [3]
                ws_new.cell(i + 2, 5, data[i][4])  # [4]
                ws_new.cell(i + 2, 6, data[i][5])  # [5]

            wb_new.save(dir + "/" + file)
        finally:
            print("已存储第{}页的数据".format(num+1))


def main():
    '''
    
    :return: None
    '''
    # 要爬取的贴吧(自行输入想要爬取的吧名)
    word = "汽车"
    kw = word.encode("UTF-8")  # 将关键词转换为Bytes格式，URL不支持中文格式
    # 贴吧页数（根据贴吧的页数自行改变）
    page = 100
    # 基础网址
    baseurl = "https://tieba.baidu.com/"
    # 将元素保存至item中
    items = []

    # 获取时间，用于获取当天的日期并返回年月日
    local_time = time.localtime(time.time())
    year = time.strftime("%Y",local_time)
    month = time.strftime("%m",local_time)
    day = time.strftime("%d",local_time)

    tb = Spider()

    for i in range(page + 1):
        # url拼接，由于搜索关键词是中文，所以需要进行处理
        url = baseurl+"f?kw="+parse.quote(kw) + "&ie=utf-8&pn=" + str(i * 50)
        html = tb.getContent(url)
        
        # 储存帖子信息
        data = tb.getInfo(html,items,year,month,day)
        
        # 将数据储存
        tb.saveData(data, word, i)

        # 每次爬取完停止数秒，防止由于爬取速度过快被限制访问
        time.sleep(3)


if __name__ == "__main__":
    main()